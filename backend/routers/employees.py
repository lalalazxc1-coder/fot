from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import select
from typing import List, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from utils.date_utils import to_iso_utc

from database.database import get_db
from database.models import User, AuditLog, Employee, OrganizationUnit, Position, FinancialRecord
from schemas import EmployeeCreate, FinancialUpdate, EmployeeUpdate, EmpDetailsUpdate
from dependencies import get_current_active_user, get_user_scope, PermissionChecker
from services.employee_service import EmployeeService

router = APIRouter(prefix="/api", tags=["employees"])

@router.post("/employees", dependencies=[Depends(PermissionChecker('edit_employees'))])
def create_employee(
    data: EmployeeCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    scope: Optional[List[int]] = Depends(get_user_scope)
):
    return EmployeeService.create_employee(db, current_user, data, scope)

@router.put("/employees/{emp_id}", dependencies=[Depends(PermissionChecker('edit_employees'))])
def update_employee(
    emp_id: int,
    data: EmployeeUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    scope: Optional[List[int]] = Depends(get_user_scope)
):
    return EmployeeService.update_employee(db, current_user, emp_id, data, scope)

from schemas import EmployeeCreate, FinancialUpdate, EmployeeUpdate, EmpDetailsUpdate, DismissEmployeeRequest

@router.post("/employees/{emp_id}/dismiss", dependencies=[Depends(PermissionChecker('edit_employees'))])
def dismiss_employee(
    emp_id: int,
    data: DismissEmployeeRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    scope: Optional[List[int]] = Depends(get_user_scope)
):
    return EmployeeService.dismiss_employee(db, current_user, emp_id, data.reason, data.date, scope)


@router.get("/employees")
def get_employees(
    db: Session = Depends(get_db), 
    current_user: User = Depends(get_current_active_user),
    scope: Optional[List[int]] = Depends(get_user_scope),
    q: Optional[str] = None
):
    return EmployeeService.get_employees(db, current_user, scope, q)

# ... (Skipping standard CRUD for brevity, focus on Export Optimization) ...

from pydantic import BaseModel
class ExportRequest(BaseModel):
    ids: Optional[List[int]] = None

@router.post("/employees/export", dependencies=[Depends(PermissionChecker('edit_employees'))])
def export_employees_excel(
    req: ExportRequest,
    db: Session = Depends(get_db), 
    current_user: User = Depends(get_current_active_user),
    scope: Optional[List[int]] = Depends(get_user_scope)
):
    # Optimized Export: Use joinedload/selectinload to fetch relations in single query
    # instead of inside the loop relying on lazy loading.
    
    # We will build query manually here for max performance
    query = select(Employee).options(
        joinedload(Employee.position), 
        joinedload(Employee.org_unit),
        # Optimizing financial record fetching is tricky with joinedload on 'latest' 
        # usually requires selectinload on collection + logic, or a subquery join.
        # But standard relationship is 'financial_records'.
    ).filter(Employee.status != "Dismissed")
    
    if scope:
        query = query.filter(Employee.org_unit_id.in_(scope))
        
    if req.ids is not None:
        query = query.filter(Employee.id.in_(req.ids))
        
    employees = db.scalars(query).all()
    
    # Batch fetch latest financials to avoid N+1
    # Strategy: Fetch all financial records for these employees, group by emp_id, take latest.
    emp_ids = [e.id for e in employees]
    financials_map = {}
    
    if emp_ids:
        # Window function or simply fetching all and sorting in python (if dataset not huge)
        # For huge datasets, subquery is better.
        # Subquery for max id per employee
        from sqlalchemy import func
        subq = select(
            FinancialRecord.employee_id, 
            func.max(FinancialRecord.id).label('max_id')
        ).where(FinancialRecord.employee_id.in_(emp_ids)).group_by(FinancialRecord.employee_id).subquery()
        
        # Join to get full record
        stmt = select(FinancialRecord).join(
            subq, 
            (FinancialRecord.employee_id == subq.c.employee_id) & 
            (FinancialRecord.id == subq.c.max_id)
        )
        fin_records = db.scalars(stmt).all()
        financials_map = {r.employee_id: r for r in fin_records}

    # Pre-fetch unit map
    org_units = db.scalars(select(OrganizationUnit)).all()
    unit_map = {u.id: u for u in org_units}

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Сотрудники"
    
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    headers = [
        "ФИО", "Должность", "Филиал", "Отдел", "Статус",
        "Оклад (Нет)", "Оклад (Брут)", "KPI (Нет)", "KPI (Брут)",
        "Бонус (Нет)", "Бонус (Брут)", "Всего (Нет)", "Всего (Брут)"
    ]
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, emp in enumerate(employees, 2):
        fin = financials_map.get(emp.id)
        
        branch_name = "-"
        dept_name = "-"
        
        if emp.org_unit_id and emp.org_unit_id in unit_map:
            u = unit_map[emp.org_unit_id]
            if u.type in ['branch', 'head_office']:
                branch_name = u.name
            else:
                dept_name = u.name
                p = unit_map.get(u.parent_id)
                while p:
                    if p.type in ['branch', 'head_office']:
                        branch_name = p.name
                        break
                    p = unit_map.get(p.parent_id)
        
        # Financials
        b_n = fin.base_net if fin else 0
        b_g = fin.base_gross if fin else 0
        k_n = fin.kpi_net if fin else 0
        k_g = fin.kpi_gross if fin else 0
        bo_n = fin.bonus_net if fin else 0
        bo_g = fin.bonus_gross if fin else 0
        t_n = fin.total_net if fin else 0
        t_g = fin.total_gross if fin else 0

        row_data = [
            emp.full_name,
            emp.position.title if emp.position else "-",
            branch_name, # Simplified for now, ideally use map
            dept_name,
            emp.status,
            b_n, b_g, k_n, k_g, bo_n, bo_g, t_n, t_g
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx >= 6: 
                cell.number_format = '#,##0'
                
    # Save
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = f"Employees_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=out.getvalue(), 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@router.get("/audit-logs/{emp_id}")
def get_employee_history(
    emp_id: int, 
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    scope_ids: list = Depends(get_user_scope)
):
    """
    Returns the audit log for a specific employee, formatted for the frontend HistoryModal.
    FIX #19: Now checks user scope before returning data.
    """
    # Scope check
    emp = db.query(Employee).filter(Employee.id == emp_id).first()
    if not emp:
        raise HTTPException(404, "Employee not found")
    if scope_ids is not None and emp.org_unit_id not in scope_ids:
        raise HTTPException(403, "Access denied to this employee's audit logs")

    logs = (
        db.query(AuditLog)
        .filter(AuditLog.target_entity == "employee", AuditLog.target_entity_id == emp_id)
        .order_by(AuditLog.id.desc())
        .all()
    )
    
    # Pre-fetch users to avoid N+1
    user_ids = set(l.user_id for l in logs)
    users = db.query(User).filter(User.id.in_(user_ids)).all()
    user_map = {u.id: u.full_name or u.email for u in users}
    
    formatted_logs = []
    for log in logs:
        # Each log can have multiple changed fields in old_values/new_values
        keys = set()
        if log.new_values: keys.update(log.new_values.keys())
        if log.old_values: keys.update(log.old_values.keys())
        
        user_name = user_map.get(log.user_id, "Система")
        
        # Sort keys to make 'created' first if present
        sorted_keys = sorted(list(keys), key=lambda x: 0 if x == 'created' else 1)
        
        for k in sorted_keys:
             formatted_logs.append({
                "date": to_iso_utc(log.timestamp) or log.timestamp,
                "user": user_name,
                "field": k,
                "oldVal": str(log.old_values.get(k, '') if log.old_values else ''),
                "newVal": str(log.new_values.get(k, '') if log.new_values else '')
            })
            
    return formatted_logs
