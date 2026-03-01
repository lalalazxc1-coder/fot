from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy.orm import Session
from sqlalchemy import desc
from typing import List, Optional
from datetime import datetime
from database.database import get_db
from database.models import PlanningPosition, User, OrganizationUnit, AuditLog
from dependencies import get_current_active_user, get_user_scope
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import logging

logger = logging.getLogger("fot.planning")

router = APIRouter(prefix="/api", tags=["planning"])

# --- Schemas ---
class PlanCreate(BaseModel):
    position: str
    branch_id: int
    department_id: int | None = None
    schedule: str | None = None
    count: int = 1
    
    base_net: int = 0
    base_gross: int = 0
    kpi_net: int = 0
    kpi_gross: int = 0
    bonus_net: int = 0
    bonus_gross: int = 0
    bonus_count: int | None = None

class PlanUpdate(BaseModel):
    position: str | None = None
    branch_id: int | None = None
    department_id: int | None = None
    schedule: str | None = None
    count: int | None = None
    
    base_net: int | None = None
    base_gross: int | None = None
    kpi_net: int | None = None
    kpi_gross: int | None = None
    bonus_net: int | None = None
    bonus_gross: int | None = None
    bonus_count: int | None = None

# --- Helpers ---
def check_manage_permission(user: User):
    if user.role_rel and user.role_rel.permissions.get('admin_access'): return True
    return user.role_rel.permissions.get('manage_planning', False) if user.role_rel and user.role_rel.permissions else False

def filter_by_scope(query, user: User, db: Session):
    from dependencies import get_user_scope
    from sqlalchemy import or_

    allowed_ids = get_user_scope(db, user)
    
    # Non-admin with scope restrictions
    if allowed_ids is not None:
        # User only sees planning positions if they fall exactly into the allowed IDs list.
        # This means EITHER it's attached directly to an allowed branch OR an allowed department
        return query.filter(
            or_(
                PlanningPosition.branch_id.in_(allowed_ids),
                PlanningPosition.department_id.in_(allowed_ids)
            )
        )
    
    # If no restrictions (or admin), return all
    return query


# --- Endpoints ---

@router.get("/planning")
def get_planning(db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    # Everyone with access to system can view? Or restricted?
    # Let's apply scope filtering.
    q = db.query(PlanningPosition).filter(PlanningPosition.scenario_id == None)
    q = filter_by_scope(q, current_user, db)
    plans = q.all()
    
    # Format for frontend
    res = []
    for p in plans:
        res.append({
            "id": p.id,
            "position": p.position_title,
            "branch_id": p.branch_id,
            "department_id": p.department_id,
            "schedule": p.schedule,
            "count": p.count,
            "base_net": p.base_net,
            "base_gross": p.base_gross,
            "kpi_net": p.kpi_net,
            "kpi_gross": p.kpi_gross,
            "bonus_net": p.bonus_net,
            "bonus_gross": p.bonus_gross,
            "bonus_count": p.bonus_count
        })
    return res

@router.post("/planning")
def create_plan(plan: PlanCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    if not check_manage_permission(current_user):
        raise HTTPException(403, "Permission 'manage_planning' required")

    # FIX #M2: Используем единый get_user_scope() вместо ручной проверки по scope_branches
    allowed_ids = get_user_scope(db, current_user)
    if allowed_ids is not None:
        # Пользователь видит только определённые единицы
        if plan.branch_id not in allowed_ids and (plan.department_id is None or plan.department_id not in allowed_ids):
            raise HTTPException(403, "Cannot create plan for this branch/department (Out of scope)")

    # Validation
    if not plan.schedule:
        raise HTTPException(400, "Необходимо указать график работы")
        
    if plan.base_net <= 0 and plan.base_gross <= 0:
        raise HTTPException(400, "Необходимо указать оклад (Net или Gross)")

    new_plan = PlanningPosition(
        position_title=plan.position,
        branch_id=plan.branch_id,
        department_id=plan.department_id,
        schedule=plan.schedule,
        count=plan.count,
        base_net=plan.base_net,
        base_gross=plan.base_gross,
        kpi_net=plan.kpi_net,
        kpi_gross=plan.kpi_gross,
        bonus_net=plan.bonus_net,
        bonus_gross=plan.bonus_gross,
        bonus_count=plan.bonus_count
    )
    db.add(new_plan)
    db.commit()
    db.refresh(new_plan)
    
    # CLEANUP: Remove old logs if ID was reused (SQLite quirk)
    db.query(AuditLog).filter_by(target_entity="planning", target_entity_id=new_plan.id).delete()
    
    # Audit
    ts = datetime.now().strftime("%d.%m.%Y %H:%M")
    audit = AuditLog(
        user_id=current_user.id,
        target_entity="planning",
        target_entity_id=new_plan.id,
        timestamp=ts,
        old_values=None,
        new_values={
            "Событие": "Создана позиция",
            "Должность": plan.position,
            "Филиал (ID)": plan.branch_id,
            "Отдел (ID)": plan.department_id,
            "График": plan.schedule,
            "Количество": plan.count,
            "Оклад (Net)": plan.base_net,
            "Оклад (Gross)": plan.base_gross,
            "KPI (Net)": plan.kpi_net,
            "KPI (Gross)": plan.kpi_gross,
            "Доплаты (Net)": plan.bonus_net,
            "Доплаты (Gross)": plan.bonus_gross,
            "Кол-во получателей доплаты": plan.bonus_count
        }
    )
    db.add(audit)
    db.commit()
    
    return {"status": "success", "id": new_plan.id}

# --- Private Helpers ---

def _sync_employee_financials(db: Session, plan: PlanningPosition, changes: dict, user: User, audit_ts: str):
    from services.salary_service import sync_employee_financials
    return sync_employee_financials(db, plan, changes, user, audit_ts)


@router.patch("/planning/{plan_id}")
def update_plan(plan_id: int, plan: PlanUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    if not check_manage_permission(current_user):
        raise HTTPException(403, "Permission 'manage_planning' required")
        
    db_plan = db.get(PlanningPosition, plan_id)
    if not db_plan: raise HTTPException(404, "Plan not found")
    
    # FIX #M2: Scope Check — единая проверка через get_user_scope()
    allowed_ids = get_user_scope(db, current_user)
    if allowed_ids is not None:
        if db_plan.branch_id not in allowed_ids and (db_plan.department_id is None or db_plan.department_id not in allowed_ids):
            raise HTTPException(403, "Out of scope")

    changes = {}
    update_data = plan.model_dump(exclude_unset=True)
    
    for key, val in update_data.items():
        db_key = key
        # Handle field mapping
        if key == 'position': db_key = 'position_title'
        
        old_val = getattr(db_plan, db_key)
        if old_val != val:
            changes[key] = {'old': old_val, 'new': val}
            setattr(db_plan, db_key, val)
            
    if changes:
        ts = datetime.now().strftime("%d.%m.%Y %H:%M")
        audit = AuditLog(
            user_id=current_user.id,
            target_entity="planning",
            target_entity_id=plan_id,
            timestamp=ts,
            old_values={k: v['old'] for k,v in changes.items()},
            new_values={k: v['new'] for k,v in changes.items()}
        )
        db.add(audit)
        
        # Auto-sync logic moved to helper
        synced = _sync_employee_financials(db, db_plan, changes, current_user, ts)
        if synced > 0:
            logger.info("Auto-synced %d employees for plan %d", synced, plan_id)

        db.commit()
        
    return {"status": "updated"}

@router.delete("/planning/{plan_id}")
def delete_plan(plan_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    if not check_manage_permission(current_user):
        raise HTTPException(403, "Permission 'manage_planning' required")

    db_plan = db.get(PlanningPosition, plan_id)
    if not db_plan: raise HTTPException(404, "Plan not found")
    
    # FIX #M2: Scope Check — единая проверка через get_user_scope()
    allowed_ids = get_user_scope(db, current_user)
    if allowed_ids is not None:
        if db_plan.branch_id not in allowed_ids and (db_plan.department_id is None or db_plan.department_id not in allowed_ids):
            raise HTTPException(403, "Out of scope")
             
    # Delete associated history
    db.query(AuditLog).filter_by(target_entity="planning", target_entity_id=plan_id).delete()
    
    db.delete(db_plan)
    db.commit()
    
    return {"status": "deleted"}

@router.get("/planning/{plan_id}/history")
def get_plan_history(plan_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    logs = db.query(AuditLog).filter_by(target_entity_id=plan_id, target_entity="planning").all()
    logs.reverse()
    formatted_logs = []
    
    user_ids = set(l.user_id for l in logs)
    users = db.query(User).filter(User.id.in_(user_ids)).all()
    user_map = {u.id: u.full_name for u in users}

    for log in logs:
        # Determine fields. For creation new_values has all. For update only changes.
        keys = set()
        if log.new_values: keys.update(log.new_values.keys())
        if log.old_values: keys.update(log.old_values.keys())
        
        for k in keys:
             formatted_logs.append({
                "date": log.timestamp,
                "user": user_map.get(log.user_id, "Unknown"),
                "field": k,
                "oldVal": str(log.old_values.get(k, '') if log.old_values else ''),
                "newVal": str(log.new_values.get(k, '') if log.new_values else '')
            })
            
    return formatted_logs


class ExportRequest(BaseModel):
    ids: Optional[List[int]] = None

@router.post("/planning/export")
def export_planning_excel(req: ExportRequest, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    perms = current_user.role_rel.permissions if current_user.role_rel else {}
    can_export = bool(perms.get("admin_access") or perms.get("manage_planning") or perms.get("view_financial_reports"))
    if not can_export:
        raise HTTPException(403, "Permission 'manage_planning' required")

    try:
        # Optimized: Pre-fetch Organization Units to avoid N+1 queries
        from sqlalchemy import select
        units = db.scalars(select(OrganizationUnit)).all()
        unit_map = {u.id: u for u in units}
        
        # Get planning data with scope filtering
        query = db.query(PlanningPosition).filter(PlanningPosition.scenario_id == None)
        query = filter_by_scope(query, current_user, db)
        if req.ids is not None:
            query = query.filter(PlanningPosition.id.in_(req.ids))
        plans = query.all()
        
        logger.info("Found %d planning positions to export", len(plans))
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "ФОТ Планирование"
        
        # Header styling
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        # Headers
        headers = [
            "Позиция", "Филиал", "Отдел", "График", "Кол-во",
            "Оклад (Нет)", "Оклад (Брут)", "KPI (Нет)", "KPI (Брут)",
            "Бонус (Нет)", "Бонус (Брут)", "Кол-во доплат", "Всего (Нет)", "Всего (Брут)"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for row_num, plan in enumerate(plans, 2):
            branch_name = "-"
            dept_name = "-"
            
            # Lookup from pre-fetched map (O(1)) instead of DB query (O(N))
            if plan.branch_id:
                branch = unit_map.get(plan.branch_id)
                if branch: branch_name = branch.name
            
            if plan.department_id:
                dept = unit_map.get(plan.department_id)
                if dept: dept_name = dept.name
            
            actual_bonus_count = plan.bonus_count if plan.bonus_count is not None else plan.count
            total_net = (plan.base_net + plan.kpi_net) * plan.count + (plan.bonus_net * actual_bonus_count)
            total_gross = (plan.base_gross + plan.kpi_gross) * plan.count + (plan.bonus_gross * actual_bonus_count)
            
            data = [
                plan.position_title,
                branch_name,
                dept_name,
                plan.schedule or "-",
                plan.count,
                plan.base_net,
                plan.base_gross,
                plan.kpi_net,
                plan.kpi_gross,
                plan.bonus_net,
                plan.bonus_gross,
                actual_bonus_count,
                total_net,
                total_gross
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if col_num >= 5:  # Number columns
                    cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="left" if col_num <= 4 else "right")
        
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws.column_dimensions[col].width = 14
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_content = excel_file.getvalue()
        
        logger.info("Excel file size: %d bytes", len(excel_content))
        
        filename = f"FOT_Planning_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return Response(
            content=excel_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception:
        logger.exception("Error exporting planning")
        raise HTTPException(status_code=500, detail="Export failed. Please try again later.")
